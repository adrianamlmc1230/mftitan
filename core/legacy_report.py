"""舊版相容報表匯出模組。

產生與舊版 gen_report.py 相同結構的 Excel 報表：
- 按洲別（continent）分 sheet
- 每個聯賽一個區塊，RT 和 Early 並排
- 每個區塊包含最多 2 個分組的 5 大區間 Home/Away 訊號
- 輸出 2 個檔案：讓球Report.xlsx（HDP）和 大小Report.xlsx（OU）
"""

from __future__ import annotations

import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 5 大區間標籤（與舊版 MST 檔案 U 欄一致）
ZONE_LABELS = [
    "*-50~-24%   ",
    "*-23~-8%     ",
    "*-7~+7%       ",
    "*+8~+23%   ",
    "*+24~+50%    ",
]


def generate_legacy_report(
    decisions: list[dict],
    leagues: dict[int, "League"],
    group_ids: list[int],
    group_names: list[str],
    play_type: str,
) -> bytes:
    """產生舊版格式的 Excel 報表。

    Args:
        decisions: decision_results 列表（已篩選 play_type）
        leagues: league_id → League 對照表
        group_ids: 要匯出的全域分組 ID 列表（最多 2 個）
        group_names: 對應的分組顯示名稱列表
        play_type: 'HDP' 或 'OU'

    Returns:
        Excel 檔案的 bytes
    """
    # Index decisions by (league_id, global_group_id, timing)
    idx: dict[tuple[int, int, str], dict] = {}
    for d in decisions:
        key = (d["league_id"], d.get("global_group_id", 0), d["timing"])
        idx[key] = d

    # Collect all league_ids that have data for any of the selected groups
    league_ids_with_data: set[int] = set()
    for d in decisions:
        if d.get("global_group_id") in group_ids:
            league_ids_with_data.add(d["league_id"])

    # Group leagues by continent
    by_continent: dict[str, list] = {}
    for lid in sorted(league_ids_with_data):
        lg = leagues.get(lid)
        if not lg:
            continue
        cont = lg.continent or "OTHER"
        by_continent.setdefault(cont, []).append(lg)

    # Sort continents and leagues within each
    for cont in by_continent:
        by_continent[cont].sort(key=lambda lg: lg.code)

    wb = Workbook()
    sorted_continents = sorted(by_continent.keys())

    # Styles
    region_font = Font(size=14, bold=True, color="FFFFFF")
    region_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    loc_font = Font(bold=True, size=12)
    rt_font = Font(bold=True, color="0000FF")
    early_font = Font(bold=True, color="FF0000")
    header_font = Font(bold=True)

    for i, continent in enumerate(sorted_continents):
        if i == 0:
            ws = wb.active
            ws.title = continent
        else:
            ws = wb.create_sheet(title=continent)

        # Column widths
        ws.column_dimensions["A"].width = 30
        for col_letter in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 14
        ws.column_dimensions["G"].width = 2
        for col_letter in ["H", "I", "J", "K", "L"]:
            ws.column_dimensions[col_letter].width = 14

        current_row = 1

        # Region header
        cell = ws.cell(row=current_row, column=1, value=f"Region: {continent}")
        cell.font = region_font
        cell.fill = region_fill
        current_row += 1

        for lg in by_continent[continent]:
            # Build location display name
            # Extract short name from name_zh (remove continent prefix if present)
            short_name = lg.name_zh
            phase_suffix = f"（{lg.phase}）" if lg.phase else ""
            loc_display = f"Location: {lg.code} ({short_name}{phase_suffix})"

            # Location header row
            ws.cell(row=current_row, column=1, value=loc_display).font = loc_font
            ws.cell(row=current_row, column=2, value="Type: RT").font = rt_font
            ws.cell(row=current_row, column=8, value="Type: Early").font = early_font
            current_row += 1

            # Write data blocks for each selected group
            for g_idx, (gid, gname) in enumerate(zip(group_ids, group_names)):
                # Group sub-header
                ws.cell(row=current_row, column=2, value=gname).font = header_font
                ws.cell(row=current_row, column=3, value="主場").font = header_font
                # D is empty
                ws.cell(row=current_row, column=5, value=gname).font = header_font
                ws.cell(row=current_row, column=6, value="作客").font = header_font

                ws.cell(row=current_row, column=8, value=gname).font = header_font
                ws.cell(row=current_row, column=9, value="主場").font = header_font
                # J is empty
                ws.cell(row=current_row, column=11, value=gname).font = header_font
                ws.cell(row=current_row, column=12, value="作客").font = header_font
                current_row += 1

                # 5 zone rows
                d_rt = idx.get((lg.id, gid, "RT"))
                d_early = idx.get((lg.id, gid, "Early"))

                for z in range(5):
                    # RT side (B-F)
                    ws.cell(row=current_row, column=2, value=ZONE_LABELS[z])
                    if d_rt:
                        h_sig = d_rt["home_signals"][z] if z < len(d_rt["home_signals"]) else ""
                        a_sig = d_rt["away_signals"][z] if z < len(d_rt["away_signals"]) else ""
                        ws.cell(row=current_row, column=3, value=h_sig)
                        # D is empty
                        ws.cell(row=current_row, column=5, value=ZONE_LABELS[z])
                        ws.cell(row=current_row, column=6, value=a_sig)
                    else:
                        ws.cell(row=current_row, column=5, value=ZONE_LABELS[z])

                    # Early side (H-L)
                    ws.cell(row=current_row, column=8, value=ZONE_LABELS[z])
                    if d_early:
                        h_sig = d_early["home_signals"][z] if z < len(d_early["home_signals"]) else ""
                        a_sig = d_early["away_signals"][z] if z < len(d_early["away_signals"]) else ""
                        ws.cell(row=current_row, column=9, value=h_sig)
                        # J is empty
                        ws.cell(row=current_row, column=11, value=ZONE_LABELS[z])
                        ws.cell(row=current_row, column=12, value=a_sig)
                    else:
                        ws.cell(row=current_row, column=11, value=ZONE_LABELS[z])

                    current_row += 1

            # Blank separator rows between leagues
            current_row += 2

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
