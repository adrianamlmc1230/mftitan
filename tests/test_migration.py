"""Tests for Directory.xlsx migration utility."""

from __future__ import annotations

import os
import tempfile

import openpyxl
import pytest

from core.config_store import ConfigStore
from utils.migration import import_directory, import_previous_directory


@pytest.fixture
def store():
    """Create a temporary ConfigStore for testing."""
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    s = ConfigStore(db_path=path)
    s.init_db()
    yield s
    # Close SQLite connection before deleting on Windows
    s._conn.close()
    try:
        os.unlink(path)
    except PermissionError:
        pass  # Windows file locking; temp file will be cleaned up later


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a minimal Directory.xlsx for testing."""
    filepath = str(tmp_path / "Directory.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2025"

    # Row 2: continent header in B (merged in real file, but value in B is enough)
    ws.cell(row=2, column=2, value="亞洲(ASI)")

    # Row 4: 前/尾 headers
    ws.cell(row=4, column=3, value="前")
    ws.cell(row=4, column=4, value="尾")
    ws.cell(row=4, column=7, value="前")
    ws.cell(row=4, column=8, value="尾")

    # League 1: cols B-D
    ws.cell(row=5, column=2, value="澳超")  # name_zh
    ws.cell(row=6, column=2, value="AUS1")  # code
    ws.cell(row=5, column=3, value="墨爾本城")  # Top team
    ws.cell(row=6, column=3, value="悉尼FC")  # Top team
    ws.cell(row=5, column=4, value="西悉尼流浪者")  # Weak team
    ws.cell(row=6, column=4, value="紐卡素噴射機")  # Weak team

    # League 2: cols F-H
    ws.cell(row=5, column=6, value="印尼超")  # name_zh
    ws.cell(row=6, column=6, value="IDN1")  # code
    ws.cell(row=5, column=7, value="佩西加雅加達")  # Top team
    ws.cell(row=5, column=8, value="傑帕拉")  # Weak team
    ws.cell(row=6, column=8, value="馬都拉聯")  # Weak team

    # Row 14: next continent (to define section boundary)
    ws.cell(row=14, column=2, value="非洲(AFR)")

    # EUR section
    ws.cell(row=14, column=2, value="非洲(AFR)")
    ws.cell(row=16, column=3, value="前")
    ws.cell(row=16, column=4, value="尾")
    ws.cell(row=17, column=2, value="南非超")
    ws.cell(row=18, column=2, value="RSA1")
    ws.cell(row=17, column=3, value="奧蘭多海盜")
    ws.cell(row=17, column=4, value="茨普帕")

    wb.save(filepath)
    wb.close()
    return filepath


def test_import_creates_leagues(store, sample_xlsx):
    """Test that import creates leagues with correct data."""
    result = import_directory(store, sample_xlsx)

    assert result["created"] == 3
    assert result["skipped"] == 0
    assert result["errors"] == []
    assert result["total_teams"] > 0

    leagues = store.list_leagues(active_only=False)
    codes = {lg.code for lg in leagues}
    assert "AUS1" in codes
    assert "IDN1" in codes
    assert "RSA1" in codes


def test_import_creates_season_and_groups(store, sample_xlsx):
    """Test that each league gets a current season with Top/Weak groups."""
    import_directory(store, sample_xlsx)

    leagues = store.list_leagues(active_only=False)
    for league in leagues:
        seasons = store.list_season_instances(league.id)
        assert len(seasons) == 1
        assert seasons[0].role == "current"

        groups = store.list_team_groups(seasons[0].id)
        group_names = {g.name for g in groups}
        assert "Top" in group_names
        assert "Weak" in group_names


def test_import_sets_teams(store, sample_xlsx):
    """Test that teams are correctly assigned to groups."""
    import_directory(store, sample_xlsx)

    leagues = store.list_leagues(active_only=False)
    aus = next(lg for lg in leagues if lg.code == "AUS1")
    seasons = store.list_season_instances(aus.id)
    groups = store.list_team_groups(seasons[0].id)

    top_group = next(g for g in groups if g.name == "Top")
    weak_group = next(g for g in groups if g.name == "Weak")

    top_teams = store.list_teams(top_group.id)
    weak_teams = store.list_teams(weak_group.id)

    assert "墨爾本城" in top_teams
    assert "悉尼FC" in top_teams
    assert "西悉尼流浪者" in weak_teams
    assert "紐卡素噴射機" in weak_teams


def test_import_idempotent(store, sample_xlsx):
    """Test that importing twice skips existing leagues (Req 1.4)."""
    result1 = import_directory(store, sample_xlsx)
    assert result1["created"] == 3

    result2 = import_directory(store, sample_xlsx)
    assert result2["created"] == 0
    assert result2["skipped"] == 3

    leagues = store.list_leagues(active_only=False)
    assert len(leagues) == 3


def test_import_continent_assignment(store, sample_xlsx):
    """Test that continents are correctly assigned."""
    import_directory(store, sample_xlsx)

    leagues = store.list_leagues(active_only=False)
    aus = next(lg for lg in leagues if lg.code == "AUS1")
    rsa = next(lg for lg in leagues if lg.code == "RSA1")

    assert aus.continent == "ASI"
    assert rsa.continent == "AFR"


def test_import_file_not_found(store):
    """Test graceful handling of missing file."""
    result = import_directory(store, "/nonexistent/Directory.xlsx")
    assert result["created"] == 0
    assert len(result["errors"]) > 0


def test_import_with_real_directory_xlsx(store):
    """Integration test with the actual Directory.xlsx file."""
    real_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "file", "Directory.xlsx"
    )
    if not os.path.exists(real_path):
        pytest.skip("Real Directory.xlsx not found")

    result = import_directory(store, real_path)

    # Should have created many leagues with no errors
    assert result["created"] > 0
    assert result["total_teams"] > 0

    leagues = store.list_leagues(active_only=False)
    codes = {lg.code for lg in leagues}

    # Verify some known leagues exist
    assert "ENG1" in codes
    assert "AUS1" in codes
    assert "BRA1" in codes

    # Verify continent assignments
    eng = next(lg for lg in leagues if lg.code == "ENG1")
    assert eng.continent == "EUR"

    aus = next(lg for lg in leagues if lg.code == "AUS1")
    assert aus.continent == "ASI"

    bra = next(lg for lg in leagues if lg.code == "BRA1")
    assert bra.continent == "AME"

    # Verify each league has a current season with Top/Weak groups
    for league in leagues:
        seasons = store.list_season_instances(league.id)
        assert len(seasons) == 1, f"League {league.code} should have 1 season"
        assert seasons[0].role == "current"

        groups = store.list_team_groups(seasons[0].id)
        group_names = {g.name for g in groups}
        assert "Top" in group_names, f"League {league.code} missing Top group"
        assert "Weak" in group_names, f"League {league.code} missing Weak group"

    # Verify idempotency
    result2 = import_directory(store, real_path)
    assert result2["created"] == 0
    assert result2["skipped"] == result["created"]


# ---------------------------------------------------------------------------
# Previous-season import tests (import_previous_directory)
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_prev_xlsx(tmp_path):
    """Create a minimal Diectory2024.xlsx for testing previous season."""
    filepath = str(tmp_path / "Diectory2024.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2024"

    # Same structure as Directory.xlsx but different teams
    ws.cell(row=2, column=2, value="亞洲(ASI)")

    ws.cell(row=4, column=3, value="前")
    ws.cell(row=4, column=4, value="尾")
    ws.cell(row=4, column=7, value="前")
    ws.cell(row=4, column=8, value="尾")

    # AUS1 with different teams than current season
    ws.cell(row=5, column=2, value="澳超")
    ws.cell(row=6, column=2, value="AUS1")
    ws.cell(row=5, column=3, value="中岸水手")  # Different Top team
    ws.cell(row=6, column=3, value="阿德萊德聯")  # Different Top team
    ws.cell(row=5, column=4, value="珀斯光輝")  # Different Weak team
    ws.cell(row=6, column=4, value="紐卡素噴射機")  # Same Weak team

    # IDN1 with different teams
    ws.cell(row=5, column=6, value="印尼超")
    ws.cell(row=6, column=6, value="IDN1")
    ws.cell(row=5, column=7, value="巴厘聯")  # Different Top team
    ws.cell(row=5, column=8, value="三寶瓏")  # Different Weak team

    ws.cell(row=14, column=2, value="非洲(AFR)")
    ws.cell(row=16, column=3, value="前")
    ws.cell(row=16, column=4, value="尾")
    ws.cell(row=17, column=2, value="南非超")
    ws.cell(row=18, column=2, value="RSA1")
    ws.cell(row=17, column=3, value="凱撒酋長")  # Different Top team
    ws.cell(row=17, column=4, value="斯泰倫博斯")  # Different Weak team

    wb.save(filepath)
    wb.close()
    return filepath


def test_prev_import_creates_previous_season(store, sample_xlsx, sample_prev_xlsx):
    """Test that previous import creates previous seasons for existing leagues."""
    # First import current season
    import_directory(store, sample_xlsx)

    # Then import previous season
    result = import_previous_directory(store, sample_prev_xlsx)

    assert result["created"] == 3
    assert result["skipped"] == 0
    assert result["no_league"] == 0
    assert result["errors"] == []
    assert result["total_teams"] > 0


def test_prev_import_different_teams(store, sample_xlsx, sample_prev_xlsx):
    """Test that previous season has different teams from current season."""
    import_directory(store, sample_xlsx)
    import_previous_directory(store, sample_prev_xlsx)

    leagues = store.list_leagues(active_only=False)
    aus = next(lg for lg in leagues if lg.code == "AUS1")
    seasons = store.list_season_instances(aus.id)

    current = next(s for s in seasons if s.role == "current")
    previous = next(s for s in seasons if s.role == "previous")

    curr_groups = store.list_team_groups(current.id)
    prev_groups = store.list_team_groups(previous.id)

    curr_top = next(g for g in curr_groups if g.name == "Top")
    prev_top = next(g for g in prev_groups if g.name == "Top")

    curr_top_teams = set(store.list_teams(curr_top.id))
    prev_top_teams = set(store.list_teams(prev_top.id))

    # Teams should be different between seasons
    assert curr_top_teams != prev_top_teams
    assert "墨爾本城" in curr_top_teams
    assert "中岸水手" in prev_top_teams


def test_prev_import_requires_existing_league(store, sample_prev_xlsx):
    """Test that previous import skips leagues not yet created."""
    # Import previous without importing current first
    result = import_previous_directory(store, sample_prev_xlsx)

    assert result["created"] == 0
    assert result["no_league"] == 3  # All 3 leagues don't exist yet


def test_prev_import_idempotent(store, sample_xlsx, sample_prev_xlsx):
    """Test that importing previous twice skips already-imported seasons."""
    import_directory(store, sample_xlsx)

    result1 = import_previous_directory(store, sample_prev_xlsx)
    assert result1["created"] == 3

    result2 = import_previous_directory(store, sample_prev_xlsx)
    assert result2["created"] == 0
    assert result2["skipped"] == 3


def test_prev_import_year_label_from_sheet(store, sample_xlsx, sample_prev_xlsx):
    """Test that year label is auto-detected from sheet name."""
    import_directory(store, sample_xlsx)
    import_previous_directory(store, sample_prev_xlsx)

    leagues = store.list_leagues(active_only=False)
    aus = next(lg for lg in leagues if lg.code == "AUS1")
    seasons = store.list_season_instances(aus.id)
    previous = next(s for s in seasons if s.role == "previous")

    assert previous.label == "2024"
    assert previous.year_start == 2024


def test_prev_import_with_real_files(store):
    """Integration test with actual Directory.xlsx + Diectory2024.xlsx."""
    base = os.path.join(os.path.dirname(__file__), "..", "..", "file")
    curr_path = os.path.join(base, "Directory.xlsx")
    prev_path = os.path.join(base, "Diectory2024.xlsx")

    if not os.path.exists(curr_path) or not os.path.exists(prev_path):
        pytest.skip("Real Directory files not found")

    # Import current first
    result1 = import_directory(store, curr_path)
    assert result1["created"] > 0

    # Import previous
    result2 = import_previous_directory(store, prev_path)
    assert result2["created"] > 0
    assert result2["errors"] == []

    # Verify each league now has both current and previous seasons
    leagues = store.list_leagues(active_only=False)
    for league in leagues:
        seasons = store.list_season_instances(league.id)
        roles = {s.role for s in seasons}
        assert "current" in roles, f"{league.code} missing current season"
        assert "previous" in roles, f"{league.code} missing previous season"

        # Both seasons should have Top/Weak groups
        for season in seasons:
            groups = store.list_team_groups(season.id)
            group_names = {g.name for g in groups}
            assert "Top" in group_names, (
                f"{league.code} {season.role} missing Top group"
            )
            assert "Weak" in group_names, (
                f"{league.code} {season.role} missing Weak group"
            )
