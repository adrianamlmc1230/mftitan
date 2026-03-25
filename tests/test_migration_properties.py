"""匯入完整性與冪等性屬性測試。

# Feature: football-quant-v2-refactor, Property 1: Directory.xlsx 匯入完整性
# Feature: football-quant-v2-refactor, Property 2: 匯入冪等性
"""

from __future__ import annotations

import os
import tempfile

import openpyxl
import pytest
from hypothesis import given, settings, HealthCheck
from hypothesis import strategies as st

from core.config_store import ConfigStore
from utils.migration import import_directory, import_previous_directory


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_CONTINENTS = ["ASI", "AFR", "AME", "EUR"]
_CONTINENT_LABELS = {
    "ASI": "亞洲(ASI)", "AFR": "非洲(AFR)",
    "AME": "美洲(AME)", "EUR": "歐洲(EUR)",
}


def _make_store():
    """Create a temporary ConfigStore."""
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    s = ConfigStore(db_path=path)
    s.init_db()
    return s, path


def _close_store(store, path):
    """Close store and clean up."""
    store._conn.close()
    try:
        os.unlink(path)
    except PermissionError:
        pass


# ---------------------------------------------------------------------------
# Hypothesis strategies
# ---------------------------------------------------------------------------

@st.composite
def league_data(draw):
    """Generate a list of unique leagues with teams."""
    num_leagues = draw(st.integers(min_value=1, max_value=6))
    leagues = []
    used_codes = set()
    used_names = set()

    for i in range(num_leagues):
        continent = draw(st.sampled_from(_CONTINENTS))
        # Generate unique code
        prefix = draw(st.sampled_from(["ENG", "FRA", "GER", "ITA", "ESP", "BRA",
                                        "ARG", "JPN", "KOR", "AUS", "RSA", "NGA"]))
        suffix = str(i + 1)
        code = f"{prefix}{suffix}"
        while code in used_codes:
            code = f"{prefix}{suffix}X"
        used_codes.add(code)

        name_zh = draw(st.text(
            min_size=2, max_size=4,
            alphabet=st.characters(whitelist_categories=("L",)),
        ))
        # Ensure unique name_zh to avoid UNIQUE constraint on (name_zh, phase)
        while name_zh in used_names:
            name_zh = name_zh + draw(st.text(
                min_size=1, max_size=2,
                alphabet=st.characters(whitelist_categories=("L",)),
            ))
        used_names.add(name_zh)

        num_top = draw(st.integers(min_value=0, max_value=4))
        num_weak = draw(st.integers(min_value=0, max_value=4))

        top_teams = [
            draw(st.text(min_size=2, max_size=6,
                         alphabet=st.characters(whitelist_categories=("L",))))
            for _ in range(num_top)
        ]
        weak_teams = [
            draw(st.text(min_size=2, max_size=6,
                         alphabet=st.characters(whitelist_categories=("L",))))
            for _ in range(num_weak)
        ]

        # Deduplicate
        top_teams = list(dict.fromkeys(top_teams))
        weak_teams = list(dict.fromkeys(weak_teams))

        leagues.append({
            "continent": continent,
            "code": code,
            "name_zh": name_zh,
            "top_teams": top_teams,
            "weak_teams": weak_teams,
        })

    return leagues


def _build_xlsx(leagues: list[dict], tmp_dir: str, sheet_name: str = "2025") -> str:
    """Build a Directory-format xlsx from league data.

    Each league occupies its own column group (B-D, F-H, J-L, ...).
    All leagues are placed in a single continent section for simplicity.
    """
    filepath = os.path.join(tmp_dir, f"Dir_{sheet_name}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Group leagues by continent
    by_continent: dict[str, list[dict]] = {}
    for lg in leagues:
        by_continent.setdefault(lg["continent"], []).append(lg)

    row = 2
    col_starts = [2, 6, 10, 14, 18, 22, 26, 30, 34, 38, 42, 46, 50, 54, 58]

    for continent in _CONTINENTS:
        if continent not in by_continent:
            continue

        cont_leagues = by_continent[continent]
        ws.cell(row=row, column=2, value=_CONTINENT_LABELS[continent])
        row += 2  # skip a row

        # Place each league in its own column group
        # 前/尾 headers for all column groups
        for i, lg in enumerate(cont_leagues):
            if i >= len(col_starts):
                break
            cs = col_starts[i]
            ws.cell(row=row, column=cs + 1, value="前")
            ws.cell(row=row, column=cs + 2, value="尾")
        row += 1

        # Write league data in parallel columns
        name_row = row
        max_height = 2  # minimum 2 rows (name + code)

        for i, lg in enumerate(cont_leagues):
            if i >= len(col_starts):
                break
            cs = col_starts[i]

            ws.cell(row=name_row, column=cs, value=lg["name_zh"])
            ws.cell(row=name_row + 1, column=cs, value=lg["code"])

            for j, team in enumerate(lg["top_teams"]):
                ws.cell(row=name_row + j, column=cs + 1, value=team)
            for j, team in enumerate(lg["weak_teams"]):
                ws.cell(row=name_row + j, column=cs + 2, value=team)

            height = max(len(lg["top_teams"]), len(lg["weak_teams"]), 2)
            max_height = max(max_height, height)

        row = name_row + max_height + 2

    wb.save(filepath)
    wb.close()
    return filepath


# ---------------------------------------------------------------------------
# Property 1: Directory.xlsx 匯入完整性
# ---------------------------------------------------------------------------

# Feature: football-quant-v2-refactor, Property 1: Directory.xlsx 匯入完整性
# Validates: Requirements 1.1, 1.2, 1.3

@given(data=league_data())
@settings(
    max_examples=100,
    suppress_health_check=[HealthCheck.too_slow],
    deadline=None,
)
def test_property1_import_completeness(data):
    """匯入完整性：匯入後每個聯賽都有 current 賽季和 Top/Weak 分組。"""
    store, db_path = _make_store()
    tmp_dir = tempfile.mkdtemp()
    try:
        filepath = _build_xlsx(data, tmp_dir)
        result = import_directory(store, filepath)

        # All leagues should be created (no errors expected for valid data)
        assert result["created"] == len(data), (
            f"Expected {len(data)} created, got {result['created']}, "
            f"errors: {result['errors']}"
        )

        leagues = store.list_leagues(active_only=False)
        imported_codes = {lg.code for lg in leagues}

        for lg_data in data:
            assert lg_data["code"] in imported_codes, (
                f"League {lg_data['code']} not found after import"
            )

            league = next(lg for lg in leagues if lg.code == lg_data["code"])

            # Should have exactly 1 season with role=current
            seasons = store.list_season_instances(league.id)
            assert len(seasons) == 1
            assert seasons[0].role == "current"

            # Should have Top and Weak groups
            groups = store.list_team_groups(seasons[0].id)
            group_names = {g.name for g in groups}
            assert "Top" in group_names
            assert "Weak" in group_names

            # Team lists should match input
            top_group = next(g for g in groups if g.name == "Top")
            weak_group = next(g for g in groups if g.name == "Weak")
            actual_top = store.list_teams(top_group.id)
            actual_weak = store.list_teams(weak_group.id)

            assert set(actual_top) == set(lg_data["top_teams"])
            assert set(actual_weak) == set(lg_data["weak_teams"])
    finally:
        _close_store(store, db_path)


# ---------------------------------------------------------------------------
# Property 2: 匯入冪等性
# ---------------------------------------------------------------------------

# Feature: football-quant-v2-refactor, Property 2: 匯入冪等性
# Validates: Requirements 1.4, 1.5

@given(data=league_data())
@settings(
    max_examples=100,
    suppress_health_check=[HealthCheck.too_slow],
    deadline=None,
)
def test_property2_import_idempotency(data):
    """匯入冪等性：匯入兩次後聯賽數量不變，第二次全部跳過。"""
    store, db_path = _make_store()
    tmp_dir = tempfile.mkdtemp()
    try:
        filepath = _build_xlsx(data, tmp_dir)

        result1 = import_directory(store, filepath)
        count_after_first = len(store.list_leagues(active_only=False))

        result2 = import_directory(store, filepath)
        count_after_second = len(store.list_leagues(active_only=False))

        # League count should not change
        assert count_after_first == count_after_second

        # Second import: 0 created, all skipped
        assert result2["created"] == 0
        assert result2["skipped"] == result1["created"]
    finally:
        _close_store(store, db_path)
