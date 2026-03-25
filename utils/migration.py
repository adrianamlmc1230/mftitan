"""Directory.xlsx 匯入工具。

讀取 Directory.xlsx（本季）與 Diectory2024.xlsx（上季）中的聯賽與隊伍資料，
匯入至 ConfigStore。

- import_directory: 匯入本季資料（建立聯賽 + current 賽季 + Top/Weak 分組）
- import_previous_directory: 匯入上季資料（為已存在的聯賽建立 previous 賽季 + Top/Weak 分組）

Validates: Requirements 1.1, 1.2, 1.3, 1.4, 1.5, 1.6
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from core.config_store import ConfigStore


# Continent header patterns: "亞洲(ASI)", "非洲(AFR)", "美洲(AME)", "歐洲(EUR)"
_CONTINENT_RE = re.compile(r"\(([A-Z]{3})\)")

# League code pattern: 2-4 uppercase letters followed by digit(s) and optional letter
_CODE_RE = re.compile(r"^[A-Z]{2,4}\d[A-Z0-9]*$")

# Column groups: each league occupies 3 columns (name/code, Top, Weak).
# Groups start at columns B, F, J, N, R, V, Z, AD, AH, AL, AP, AT, AX, BB, BF
# i.e. column indices 2, 6, 10, 14, 18, 22, 26, 30, 34, 38, 42, 46, 50, 54, 58
_COL_GROUP_STARTS = list(range(2, 60, 4))


def _parse_xlsx(filepath: str) -> tuple[list[dict], list[str]]:
    """Parse a Directory xlsx file and return (league_info_list, errors).

    Shared by import_directory and import_previous_directory.
    """
    errors: list[str] = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as exc:
        return [], [f"無法開啟檔案：{exc}"]

    ws = wb.active
    if ws is None:
        wb.close()
        return [], ["工作簿中沒有工作表"]

    all_leagues: list[dict] = []
    continent_sections = _find_continent_sections(ws)
    for continent_code, start_row, end_row in continent_sections:
        leagues = _extract_leagues_from_section(ws, start_row, end_row, continent_code)
        all_leagues.extend(leagues)

    wb.close()
    return all_leagues, errors


def import_directory(config_store: ConfigStore, filepath: str) -> dict:
    """Import leagues from Directory.xlsx (current season).

    Returns:
        {"created": int, "skipped": int, "total_teams": int, "errors": list[str]}
    """
    summary: dict = {"created": 0, "skipped": 0, "total_teams": 0, "errors": []}

    all_leagues, parse_errors = _parse_xlsx(filepath)
    if parse_errors:
        summary["errors"].extend(parse_errors)
        return summary

    existing_codes = {lg.code for lg in config_store.list_leagues(active_only=False)}

    for league_info in all_leagues:
        _import_single_league(config_store, league_info, existing_codes, summary)

    return summary


def import_previous_directory(
    config_store: ConfigStore,
    filepath: str,
    year_label: str | None = None,
) -> dict:
    """Import previous-season team lists from Diectory2024.xlsx.

    For each league found in the file:
    - If the league code does NOT exist in ConfigStore → skip (league must be created first)
    - If the league already has a role='previous' season → skip
    - Otherwise → create a previous season with Top/Weak groups

    Args:
        config_store: ConfigStore instance.
        filepath: Path to the previous-season Directory xlsx.
        year_label: Season label (e.g. "2024"). Auto-detected from sheet name if None.

    Returns:
        {"created": int, "skipped": int, "total_teams": int,
         "no_league": int, "errors": list[str]}
    """
    summary: dict = {
        "created": 0, "skipped": 0, "total_teams": 0,
        "no_league": 0, "errors": [],
    }

    all_leagues, parse_errors = _parse_xlsx(filepath)
    if parse_errors:
        summary["errors"].extend(parse_errors)
        return summary

    # Detect year from sheet name if not provided
    if year_label is None:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            year_label = str(ws.title).strip() if ws else "2024"
            wb.close()
        except Exception:
            year_label = "2024"

    # Build lookup: code → League
    existing_leagues = {lg.code: lg for lg in config_store.list_leagues(active_only=False)}

    for league_info in all_leagues:
        code = league_info["code"]

        # League must already exist (created by import_directory)
        if code not in existing_leagues:
            summary["no_league"] += 1
            continue

        league = existing_leagues[code]

        # Check if previous season already exists
        seasons = config_store.list_season_instances(league.id)
        has_previous = any(s.role == "previous" for s in seasons)
        if has_previous:
            summary["skipped"] += 1
            continue

        try:
            # Determine year_start from label
            try:
                year_start = int(year_label)
            except ValueError:
                year_start = 2024

            season_id = config_store.create_season_instance(
                league_id=league.id,
                label=year_label,
                year_start=year_start,
            )
            config_store.set_season_role(season_id, "previous")

            # Create Top group
            top_teams = league_info["top_teams"]
            top_group_id = config_store.create_team_group(
                season_instance_id=season_id,
                name="Top",
            )
            if top_teams:
                config_store.set_teams(top_group_id, top_teams)

            # Create Weak group
            weak_teams = league_info["weak_teams"]
            weak_group_id = config_store.create_team_group(
                season_instance_id=season_id,
                name="Weak",
            )
            if weak_teams:
                config_store.set_teams(weak_group_id, weak_teams)

            summary["created"] += 1
            summary["total_teams"] += len(top_teams) + len(weak_teams)

        except Exception as exc:
            summary["errors"].append(f"匯入聯賽 {code} 上季資料失敗：{exc}")

    return summary


def _find_continent_sections(ws: Worksheet) -> list[tuple[str, int, int]]:
    """Scan column B for continent headers, return (code, start_row, end_row) tuples."""
    headers: list[tuple[str, int]] = []
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=2).value
        if val and isinstance(val, str):
            m = _CONTINENT_RE.search(val)
            if m and m.group(1) in ("ASI", "AFR", "AME", "EUR"):
                headers.append((m.group(1), row_idx))

    sections: list[tuple[str, int, int]] = []
    for i, (code, start) in enumerate(headers):
        end = headers[i + 1][1] - 1 if i + 1 < len(headers) else ws.max_row
        sections.append((code, start, end))
    return sections


def _extract_leagues_from_section(
    ws: Worksheet, start_row: int, end_row: int, continent_code: str
) -> list[dict]:
    """Extract all league info dicts from a continent section."""
    leagues: list[dict] = []

    for col_start in _COL_GROUP_STARTS:
        # Scan rows in this column group for league code patterns
        for row_idx in range(start_row, end_row + 1):
            val = ws.cell(row=row_idx, column=col_start).value
            if val and isinstance(val, str) and _CODE_RE.match(val.strip()):
                code = val.strip()
                # League name is in the row above, same column
                name_zh = ws.cell(row=row_idx - 1, column=col_start).value
                name_zh = str(name_zh).strip() if name_zh else code

                # Determine country from name_zh (use name_zh as country fallback)
                # country removed — name_zh is the full league identity

                # Top teams: col_start+1 (前 column), from name_row downward
                top_col = col_start + 1
                weak_col = col_start + 2
                name_row = row_idx - 1

                # Find the row range for this league's teams
                # Teams span from name_row to the next league or section end
                team_end_row = _find_team_end_row(
                    ws, col_start, row_idx + 1, end_row
                )

                top_teams = _collect_teams(ws, top_col, name_row, team_end_row)
                weak_teams = _collect_teams(ws, weak_col, name_row, team_end_row)

                leagues.append(
                    {
                        "continent": continent_code,
                        "code": code,
                        "name_zh": name_zh,
                        "top_teams": top_teams,
                        "weak_teams": weak_teams,
                    }
                )
    return leagues


def _find_team_end_row(
    ws: Worksheet, col_start: int, after_code_row: int, section_end: int
) -> int:
    """Find the last row that belongs to the current league's team block.

    Teams end when we hit a "前"/"尾" header row (next league block header)
    or the section ends.
    """
    for row_idx in range(after_code_row, section_end + 1):
        val = ws.cell(row=row_idx, column=col_start + 1).value
        if val and isinstance(val, str) and val.strip() == "前":
            return row_idx - 1
    return section_end


def _collect_teams(
    ws: Worksheet, col: int, start_row: int, end_row: int
) -> list[str]:
    """Collect non-empty, unique team names from a column range."""
    teams: list[str] = []
    seen: set[str] = set()
    for row_idx in range(start_row, end_row + 1):
        val = ws.cell(row=row_idx, column=col).value
        if val is None:
            continue
        name = str(val).strip().replace("\xa0", "")
        # Skip header labels and empty strings
        if not name or name in ("前", "尾"):
            continue
        if name not in seen:
            teams.append(name)
            seen.add(name)
    return teams


def _import_single_league(
    config_store: ConfigStore,
    league_info: dict,
    existing_codes: set[str],
    summary: dict,
) -> None:
    """Import a single league into ConfigStore."""
    code = league_info["code"]

    if code in existing_codes:
        summary["skipped"] += 1
        return

    try:
        # 1. Create league
        league_id = config_store.create_league(
            continent=league_info["continent"],
            code=code,
            name_zh=league_info["name_zh"],
        )

        # 2. Create default current season
        now = datetime.now()
        label = str(now.year)
        season_id = config_store.create_season_instance(
            league_id=league_id,
            label=label,
            year_start=now.year,
        )
        config_store.set_season_role(season_id, "current")

        # 3. Create Top team group and set teams
        top_teams = league_info["top_teams"]
        top_group_id = config_store.create_team_group(
            season_instance_id=season_id,
            name="Top",
        )
        if top_teams:
            config_store.set_teams(top_group_id, top_teams)

        # 4. Create Weak team group and set teams
        weak_teams = league_info["weak_teams"]
        weak_group_id = config_store.create_team_group(
            season_instance_id=season_id,
            name="Weak",
        )
        if weak_teams:
            config_store.set_teams(weak_group_id, weak_teams)

        summary["created"] += 1
        summary["total_teams"] += len(top_teams) + len(weak_teams)
        existing_codes.add(code)

    except Exception as exc:
        summary["errors"].append(f"匯入聯賽 {code} 失敗：{exc}")
